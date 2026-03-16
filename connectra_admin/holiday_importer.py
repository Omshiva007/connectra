from openpyxl import load_workbook
from database_admin import get_connection


def import_holiday_excel(file_path):

    wb = load_workbook(file_path)
    sheet = wb.active

    conn = get_connection()
    cursor = conn.cursor()

    # clear previous calendar
    cursor.execute("DELETE FROM holiday_calendar")

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if not row or not row[0]:
            continue

        holiday = row[0]
        date = row[1]

        # Flexible columns:
        # If only Holiday + Date are provided (like your sheet),
        # we auto-fill the rest with defaults.
        region = row[2] if len(row) > 2 else "India"
        template = row[3] if len(row) > 3 else "Holiday Greeting"
        reminder_days = row[4] if len(row) > 4 and row[4] is not None else 7
        active = row[5] if len(row) > 5 else "yes"

        active_flag = 1 if str(active).lower() == "yes" else 0

        cursor.execute(
            """
            INSERT INTO holiday_calendar
            (holiday,date,region,template,reminder_days,active)
            VALUES (?,?,?,?,?,?)
            """,
            (
                holiday,
                str(date),
                region,
                template,
                int(reminder_days),
                active_flag,
            ),
        )

    conn.commit()
    conn.close()